#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
export_website_icons.py — Export mapped game icons with metadata for website integration.

Creates a structured folder with:
  - Icon PNGs organized by type (items/, equipment/, eggs/, etc.)
  - JSON manifests linking icons to game entities
  - Metadata for each icon (ID, name, category, source sheet, etc.)

Usage:
  python export_website_icons.py [--output DIR] [--scale N]
"""

import argparse
import json
import sys
from pathlib import Path
import shutil
from PIL import Image

# Add parent to path for imports
sys.path.insert(0, str(Path(__file__).parent))

import config
from parsers.csv_parser import load_items, load_equip, load_eggs
from parsers.inf_parser import parse_img_inf


REQUESTED_ICONS_DIR = Path(r"C:\Users\anisb\OneDrive\Desktop\icons i want")
LEGACY_IMAGE_ATLAS_DIR = Path(
    r"C:\Users\anisb\OneDrive\Desktop\replit kingdom adventures - Copy\KA-Legacy-Archive\kingdom-adventures-tmp\KA_assets\image_atlas"
)

# Explicit locale pins for language-sensitive requested icons.
REQUESTED_SOURCE_OVERRIDES: dict[str, str] = {
    "restart_exploration_button.png": "game/English.lproj/restart_exploration_button.png",
    "return_button.png": "game/English.lproj/return_button.png",
    "survey_button_00.png": "game/English.lproj/survey_button_00.png",
}

# Variant-2 artifacts are still unreliable for these entries; keep primary form only.
FURNITURE_PRIMARY_ONLY_VARIANTS = {
    "chair",
    "stove",
    "restaurant shelves",
}


def apply_alpha_bleed(image: Image.Image, radius: int = 1) -> Image.Image:
    """Fill transparent-edge RGB from neighboring opaque pixels to reduce scaling halos."""
    if image.mode != "RGBA":
        image = image.convert("RGBA")

    src = image.copy()
    src_px = src.load()
    w, h = src.size
    out = src.copy()
    out_px = out.load()

    for y in range(h):
        for x in range(w):
            r, g, b, a = src_px[x, y]
            if a != 0:
                continue

            sum_r = 0
            sum_g = 0
            sum_b = 0
            count = 0
            for dy in range(-radius, radius + 1):
                for dx in range(-radius, radius + 1):
                    if dx == 0 and dy == 0:
                        continue
                    nx = x + dx
                    ny = y + dy
                    if nx < 0 or ny < 0 or nx >= w or ny >= h:
                        continue
                    nr, ng, nb, na = src_px[nx, ny]
                    if na > 0:
                        sum_r += nr
                        sum_g += ng
                        sum_b += nb
                        count += 1

            if count:
                out_px[x, y] = (sum_r // count, sum_g // count, sum_b // count, 0)

    return out


def enhance_exported_pngs(output_dir: Path) -> int:
    """Apply quality post-processing to all exported PNGs."""
    updated = 0
    for png_path in output_dir.rglob("*.png"):
        try:
            img = Image.open(png_path).convert("RGBA")
            enhanced = apply_alpha_bleed(img, radius=1)
            enhanced.save(png_path, "PNG")
            updated += 1
        except Exception:
            continue
    return updated


def export_item_icons(output_dir: Path, scale: int = 1) -> list[dict]:
    """Export item icons to items/ subfolder with metadata."""
    print("\n[Items]")
    items_dir = output_dir / "items"
    items_dir.mkdir(parents=True, exist_ok=True)
    
    items = load_items(config.CSV_ITEM)
    KA = config.KA_ASSETS_DIR
    
    exported = []
    errors = []
    
    for item in items:
        item_id = item["id"]
        icon_u = item.get("iconU")
        icon_v = item.get("iconV")
        
        # Skip items without valid icons
        if icon_u is None or icon_v is None or icon_u < 0 or icon_v < 0:
            continue
        
        # Skip hidden items (15-25)
        if 15 <= item_id <= 25:
            continue
        
        try:
            # Route to correct sheet (v13/v14 routing)
            if 0 <= item_id <= 6:
                # Material resources - direct top-row crop from material_icon.png
                sheet_path = KA / "com" / "material_icon.png"
                method = "material_top_row"
                source_img = Image.open(sheet_path).convert("RGBA")
                
                src_x = item_id * 14
                src_y = 0
                w, h = 14, 14
                
                # Center on 16x16 canvas
                icon = Image.new("RGBA", (16, 16), (0, 0, 0, 0))
                cropped = source_img.crop((src_x, src_y, src_x + w, src_y + h))
                icon.paste(cropped, (1, 1))
                sheet_name = "material_icon.png"
                
            elif 15 <= item_id <= 70:
                # Localized items - from English.lproj/icon_item2.png
                # NOTE: icon_item2 uses 32×32 cells, not 16×16!
                sheet_path = KA / "com_2" / "English.lproj" / "icon_item2.png"
                opt_path = KA / "com_2" / "icon_item2.opt"
                method = "localized_item2_opt"
                
                if not sheet_path.exists() or not opt_path.exists():
                    errors.append(f"Item {item_id}: sheet or opt missing")
                    continue
                
                # Use .opt parser
                from parsers.opt_parser import parse_opt
                opt_data = parse_opt(opt_path)
                source_img = Image.open(sheet_path).convert("RGBA")
                
                # Find sprite in opt
                sprite = next((s for s in opt_data["sprites"] if s["u"] == icon_u and s["v"] == icon_v), None)
                if not sprite or sprite["status"] != "filled":
                    errors.append(f"Item {item_id}: sprite not found in opt")
                    continue
                
                # Extract using opt coordinates onto 32×32 canvas (icon_item2 cell size)
                icon = Image.new("RGBA", (32, 32), (0, 0, 0, 0))
                cropped = source_img.crop((
                    sprite["src_x"],
                    sprite["src_y"],
                    sprite["src_x"] + sprite["w"],
                    sprite["src_y"] + sprite["h"]
                ))
                icon.paste(cropped, (sprite["dest_x"], sprite["dest_y"]))
                sheet_name = "English.lproj/icon_item2.png"
                
            else:  # 71+
                # Goods/materials - from icon_item.png
                sheet_path = KA / "com" / "icon_item.png"
                opt_path = KA / "com" / "icon_item.opt"
                method = "icon_item_opt"
                
                if not sheet_path.exists() or not opt_path.exists():
                    errors.append(f"Item {item_id}: sheet or opt missing")
                    continue
                
                from parsers.opt_parser import parse_opt
                opt_data = parse_opt(opt_path)
                source_img = Image.open(sheet_path).convert("RGBA")
                
                sprite = next((s for s in opt_data["sprites"] if s["u"] == icon_u and s["v"] == icon_v), None)
                if not sprite or sprite["status"] != "filled":
                    errors.append(f"Item {item_id}: sprite not found in opt")
                    continue
                
                icon = Image.new("RGBA", (16, 16), (0, 0, 0, 0))
                cropped = source_img.crop((
                    sprite["src_x"],
                    sprite["src_y"],
                    sprite["src_x"] + sprite["w"],
                    sprite["src_y"] + sprite["h"]
                ))
                icon.paste(cropped, (sprite["dest_x"], sprite["dest_y"]))
                sheet_name = "icon_item.png"
            
            # Scale if requested
            if scale > 1:
                icon = icon.resize((icon.width * scale, icon.height * scale), Image.Resampling.NEAREST)
            
            # Save icon
            filename = f"item_{item_id:03d}.png"
            icon.save(items_dir / filename, "PNG")
            
            exported.append({
                "id": item_id,
                "name": item["name"],
                "category": item.get("category"),
                "filename": filename,
                "method": method,
                "sheet": sheet_name,
                "iconU": icon_u,
                "iconV": icon_v,
            })
            
        except Exception as e:
            errors.append(f"Item {item_id} ({item['name']}): {e}")
    
    print(f"  Exported: {len(exported)} items")
    if errors:
        print(f"  Errors: {len(errors)}")
        for err in errors[:5]:
            print(f"    {err}")
    
    return exported


def export_equipment_icons(output_dir: Path, scale: int = 1) -> list[dict]:
    """Export equipment icons to equipment/ subfolder with metadata."""
    print("\n[Equipment]")
    equip_dir = output_dir / "equipment"
    equip_dir.mkdir(parents=True, exist_ok=True)
    
    equips = load_equip(config.CSV_EQUIP)
    KA = config.KA_ASSETS_DIR
    
    exported = []
    errors = []
    
    for equip in equips:
        equip_id = equip["id"]
        icon_u = equip.get("iconU")
        icon_v = equip.get("iconV")
        equip_type = equip.get("type")
        
        if icon_u is None or icon_v is None or icon_u < 0 or icon_v < 0:
            continue
        
        try:
            # Route to correct sheet by type
            if equip_type in range(1, 11):
                sheet_name = "icon_weapon"
            elif equip_type == 11:
                sheet_name = "icon_sheild"
            elif equip_type == 12:
                sheet_name = "icon_body"
            elif equip_type == 13:
                sheet_name = "icon_head"
            elif equip_type == 14:
                sheet_name = "icon_accessory"
            else:
                errors.append(f"Equip {equip_id}: unknown type {equip_type}")
                continue
            
            sheet_path = KA / "com" / f"{sheet_name}.png"
            opt_path = KA / "com" / f"{sheet_name}.opt"
            
            # Some sheets don't have .opt files (icon_body, icon_accessory)
            has_opt = opt_path.exists()
            
            if not sheet_path.exists():
                errors.append(f"Equip {equip_id}: sheet missing")
                continue
            
            source_img = Image.open(sheet_path).convert("RGBA")
            
            if has_opt:
                from parsers.opt_parser import parse_opt
                opt_data = parse_opt(opt_path)
                sprite = next((s for s in opt_data["sprites"] if s["u"] == icon_u and s["v"] == icon_v), None)
                
                if not sprite or sprite["status"] != "filled":
                    errors.append(f"Equip {equip_id}: sprite not found in opt")
                    continue
                
                # Extract sprite into 16×16 cell
                temp_icon = Image.new("RGBA", (16, 16), (0, 0, 0, 0))
                cropped = source_img.crop((
                    sprite["src_x"],
                    sprite["src_y"],
                    sprite["src_x"] + sprite["w"],
                    sprite["src_y"] + sprite["h"]
                ))
                temp_icon.paste(cropped, (sprite["dest_x"], sprite["dest_y"]))
                
                # Tight-crop: find content bbox and add 1px padding, discard blank canvas
                bbox = temp_icon.getbbox()
                if bbox:
                    pad = 1
                    left = max(0, bbox[0] - pad)
                    top = max(0, bbox[1] - pad)
                    right = min(temp_icon.width, bbox[2] + pad)
                    bottom = min(temp_icon.height, bbox[3] + pad)
                    icon = temp_icon.crop((left, top, right, bottom))
                else:
                    icon = temp_icon
                method = f"{sheet_name}_opt"
            else:
                # Grid fallback for sheets without .opt
                cell_w, cell_h = 16, 16
                cols = source_img.width // cell_w
                
                src_x = icon_u * cell_w
                src_y = icon_v * cell_h
                
                temp_icon = source_img.crop((src_x, src_y, src_x + cell_w, src_y + cell_h))
                
                # Tight-crop: find content bbox and add 1px padding
                bbox = temp_icon.getbbox()
                if bbox:
                    pad = 1
                    left = max(0, bbox[0] - pad)
                    top = max(0, bbox[1] - pad)
                    right = min(temp_icon.width, bbox[2] + pad)
                    bottom = min(temp_icon.height, bbox[3] + pad)
                    icon = temp_icon.crop((left, top, right, bottom))
                else:
                    icon = temp_icon
                method = f"{sheet_name}_grid"
            
            # Scale if requested
            if scale > 1:
                icon = icon.resize((icon.width * scale, icon.height * scale), Image.Resampling.NEAREST)
            
            # Save icon
            filename = f"equip_{equip_id:03d}.png"
            icon.save(equip_dir / filename, "PNG")
            
            exported.append({
                "id": equip_id,
                "name": equip["name"],
                "type": equip_type,
                "attribute": equip.get("attribute"),
                "filename": filename,
                "method": method,
                "sheet": f"{sheet_name}.png",
                "iconU": icon_u,
                "iconV": icon_v,
            })
            
        except Exception as e:
            errors.append(f"Equip {equip_id} ({equip['name']}): {e}")
    
    print(f"  Exported: {len(exported)} equipment")
    if errors:
        print(f"  Errors: {len(errors)}")
        for err in errors[:5]:
            print(f"    {err}")
    
    return exported


def export_egg_icons(output_dir: Path, scale: int = 1) -> list[dict]:
    """Export egg icons to eggs/ subfolder with metadata (UNHATCHED versions only)."""
    print("\n[Eggs]")
    eggs_dir = output_dir / "eggs"
    eggs_dir.mkdir(parents=True, exist_ok=True)
    
    eggs = load_eggs(config.CSV_EGG)
    KA = config.KA_ASSETS_DIR
    
    # Load material/img.inf for filename mapping
    img_inf_path = KA / "material" / "img.inf"
    if not img_inf_path.exists():
        print("  Error: material/img.inf not found")
        return []
    
    img_inf = parse_img_inf(img_inf_path)
    
    exported = []
    errors = []
    
    for egg in eggs:
        egg_id = egg["id"]
        image_id = egg.get("image_id")
        
        if image_id is None or image_id not in img_inf:
            errors.append(f"Egg {egg_id}: image_id {image_id} not in img.inf")
            continue
        
        try:
            egg_filename = img_inf[image_id]
            egg_png_path = KA / "material" / egg_filename
            egg_opt_path = KA / "material" / egg_filename.replace('.png', '.opt')
            
            if not egg_png_path.exists():
                errors.append(f"Egg {egg_id}: {egg_filename} not found")
                continue
            
            # Extract UNHATCHED version using .opt coordinates
            # Egg PNGs are 28×55 with two states stacked vertically:
            #   Top (0-33): Unhatched egg
            #   Bottom (33-55): Hatched egg with baby monster
            # The .opt file contains sprite at [0,0] that crops just the unhatched portion
            
            if egg_opt_path.exists():
                from parsers.opt_parser import parse_opt
                opt_data = parse_opt(egg_opt_path)
                png = Image.open(egg_png_path).convert("RGBA")
                
                # Get unhatched sprite (u=0, v=0)
                sprite = next((s for s in opt_data['sprites'] if s['u'] == 0 and s['v'] == 0 and s['status'] == 'filled'), None)
                
                if sprite:
                    # Crop unhatched portion (28×33)
                    icon = png.crop((sprite['src_x'], sprite['src_y'], 
                                   sprite['src_x'] + sprite['w'], 
                                   sprite['src_y'] + sprite['h']))
                    dimensions = f"{sprite['w']}x{sprite['h']}"
                else:
                    # Fallback: crop top portion manually
                    icon = Image.open(egg_png_path).convert("RGBA").crop((0, 0, 28, 33))
                    dimensions = "28x33"
            else:
                # Fallback without .opt: crop top 28×33 manually
                icon = Image.open(egg_png_path).convert("RGBA").crop((0, 0, 28, 33))
                dimensions = "28x33"
            
            # Scale if requested
            if scale > 1:
                icon = icon.resize((icon.width * scale, icon.height * scale), Image.Resampling.NEAREST)
            
            # Save icon
            filename = f"egg_{egg_id}.png"
            icon.save(eggs_dir / filename, "PNG")
            
            exported.append({
                "id": egg_id,
                "name": egg["name"],
                "image_id": image_id,
                "filename": filename,
                "source": egg_filename,
                "dimensions": dimensions,
                "state": "unhatched",
            })
            
        except Exception as e:
            errors.append(f"Egg {egg_id} ({egg['name']}): {e}")
    
    print(f"  Exported: {len(exported)} eggs (unhatched versions)")
    if errors:
        print(f"  Errors: {len(errors)}")
        for err in errors:
            print(f"    {err}")
    
    return exported


def export_attribute_icons(output_dir: Path, scale: int = 1) -> list[dict]:
    """Export field attribute icons to attributes/ subfolder.
    
    field_attribute_icon.png is 112×28 with TWO rows:
    - Top 14px: transparent-base variants (preferred)
    - Bottom 14px: colored tiles used only to repair clipped-looking missing pixels
    """
    print("\n[Field Attributes]")
    attr_dir = output_dir / "attributes"
    attr_dir.mkdir(parents=True, exist_ok=True)
    
    KA = config.KA_ASSETS_DIR
    attr_png = KA / "com" / "field_attribute_icon.png"
    
    if not attr_png.exists():
        print("  Error: field_attribute_icon.png not found")
        return []
    
    source_img = Image.open(attr_png).convert("RGBA")
    
    attr_names = ["Ground", "Grass", "Sand", "Rock", "Volcano", "Snow", "Swamp"]
    exported = []
    
    # Use top row as requested, with bottom-row foreground repair for missing body pixels.
    ATTRIBUTE_ICON_W = 16
    ATTRIBUTE_ICON_H = 14
    
    for i, name in enumerate(attr_names, 1):
        src_x = (i - 1) * ATTRIBUTE_ICON_W
        src_y = 0
        
        top_row = source_img.crop((src_x, 0, src_x + ATTRIBUTE_ICON_W, src_y + ATTRIBUTE_ICON_H))
        bottom_row = source_img.crop((src_x, 14, src_x + ATTRIBUTE_ICON_W, 14 + ATTRIBUTE_ICON_H))

        # Extract likely foreground from bottom row by removing border-connected background.
        fg = bottom_row.copy()
        fg_px = fg.load()
        w, h = fg.size
        visited = [[False] * w for _ in range(h)]

        def _is_bg_like(c1: tuple[int, int, int, int], c2: tuple[int, int, int, int]) -> bool:
            return (
                abs(c1[0] - c2[0]) <= 10
                and abs(c1[1] - c2[1]) <= 10
                and abs(c1[2] - c2[2]) <= 10
            )

        edge_seeds: list[tuple[int, int]] = []
        for x in range(w):
            edge_seeds.append((x, 0))
            edge_seeds.append((x, h - 1))
        for y in range(h):
            edge_seeds.append((0, y))
            edge_seeds.append((w - 1, y))

        for sx, sy in edge_seeds:
            if visited[sy][sx]:
                continue
            seed_color = fg_px[sx, sy]
            stack = [(sx, sy)]
            visited[sy][sx] = True
            while stack:
                cx, cy = stack.pop()
                cur = fg_px[cx, cy]
                fg_px[cx, cy] = (0, 0, 0, 0)
                for nx, ny in ((cx + 1, cy), (cx - 1, cy), (cx, cy + 1), (cx, cy - 1)):
                    if nx < 0 or ny < 0 or nx >= w or ny >= h or visited[ny][nx]:
                        continue
                    nxt = fg_px[nx, ny]
                    if _is_bg_like(nxt, seed_color) or _is_bg_like(nxt, cur):
                        visited[ny][nx] = True
                        stack.append((nx, ny))

        # Compose: top row stays authoritative; fill only where top is transparent.
        icon = top_row.copy()
        icon_px = icon.load()
        for y in range(h):
            for x in range(w):
                if icon_px[x, y][3] == 0 and fg_px[x, y][3] > 0:
                    icon_px[x, y] = fg_px[x, y]

        canvas_w, canvas_h = 18, 16
        canvas = Image.new("RGBA", (canvas_w, canvas_h), (0, 0, 0, 0))
        # Keep a deterministic 1px margin around the native 16x14 icon cell.
        paste_x = 1
        paste_y = 1
        canvas.paste(icon, (paste_x, paste_y))
        icon = canvas
        
        # Scale if requested
        if scale > 1:
            icon = icon.resize((icon.width * scale, icon.height * scale), Image.Resampling.NEAREST)
        
        filename = f"attribute_{i}_{name.lower()}.png"
        icon.save(attr_dir / filename, "PNG")
        
        exported.append({
            "id": i,
            "name": name,
            "filename": filename,
            "dimensions": f"{icon.width}x{icon.height}",
            "rect": f"({src_x},{src_y},{ATTRIBUTE_ICON_W},{ATTRIBUTE_ICON_H})",
        })
    
    print(f"  Exported: {len(exported)} attributes")
    return exported


def export_gender_icons(output_dir: Path, scale: int = 1) -> list[dict]:
    """Export gender icons to gender/ subfolder.
    
    gender.png is 32×20 with 2 icons side by side: male (left) and female (right).
    Each icon is 16×20.
    """
    print("\n[Gender Icons]")
    gender_dir = output_dir / "gender"
    gender_dir.mkdir(parents=True, exist_ok=True)
    
    KA = config.KA_ASSETS_DIR
    gender_png = KA / "com" / "gender.png"
    
    if not gender_png.exists():
        print("  Error: gender.png not found")
        return []
    
    source_img = Image.open(gender_png).convert("RGBA")
    
    genders = [
        {"id": 0, "name": "Male", "x": 0},
        {"id": 1, "name": "Female", "x": 16}
    ]
    exported = []
    
    GENDER_ICON_W = 16
    GENDER_ICON_H = 20
    
    for gender in genders:
        src_x = gender["x"]
        src_y = 0
        
        # Crop 16×20 region
        icon = source_img.crop((src_x, src_y, src_x + GENDER_ICON_W, src_y + GENDER_ICON_H))
        
        # Scale if requested
        if scale > 1:
            icon = icon.resize((icon.width * scale, icon.height * scale), Image.Resampling.NEAREST)
        
        filename = f"gender_{gender['id']}_{gender['name'].lower()}.png"
        icon.save(gender_dir / filename, "PNG")
        
        exported.append({
            "id": gender["id"],
            "name": gender["name"],
            "filename": filename,
            "dimensions": "16x20",
            "rect": f"({src_x},{src_y},{GENDER_ICON_W},{GENDER_ICON_H})",
        })
    
    print(f"  Exported: {len(exported)} gender icons")
    return exported


def export_furniture_icons(output_dir: Path, scale: int = 1) -> list[dict]:
    """Export furniture icons using enriched MapChip name/resFolder/pngName mapping."""
    print("\n[Furniture]")
    furniture_dir = output_dir / "furniture"
    furniture_dir.mkdir(parents=True, exist_ok=True)

    # Remove stale variant files so dropped forms do not linger across exports.
    for old_png in furniture_dir.glob("*.png"):
        try:
            old_png.unlink()
        except OSError:
            pass

    KA = config.KA_ASSETS_DIR
    furniture_assets_dir = KA / "furniture"
    if not furniture_assets_dir.exists():
        print("  Error: KA_assets/furniture not found")
        return []

    # Prefer the enriched workbook (external RE path first), fallback to workspace copy.
    enriched_candidates = [
        Path(r"C:\APK-RE\kingdom-adventurers\KA GameData - enriched asset names.xlsx"),
        config.CSV_DIR / "KA GameData - enriched asset names.xlsx",
        config.WORKSPACE_ROOT / "KA GameData - enriched asset names.xlsx",
    ]
    enriched_path = next((p for p in enriched_candidates if p.exists()), None)
    if not enriched_path:
        print("  Error: enriched workbook not found (KA GameData - enriched asset names.xlsx)")
        return []

    img_inf_path = furniture_assets_dir / "img.inf"
    if not img_inf_path.exists():
        print("  Error: furniture/img.inf not found")
        return []
    img_inf = parse_img_inf(img_inf_path)

    atlas_size_by_png: dict[str, tuple[int, int]] = {}
    try:
        from parsers.atlas_parser import parse_atlas_txt

        if LEGACY_IMAGE_ATLAS_DIR.exists():
            for atlas_txt in sorted(LEGACY_IMAGE_ATLAS_DIR.glob("ImageAtlas*.txt")):
                for sprite in parse_atlas_txt(atlas_txt):
                    atlas_size_by_png.setdefault(sprite.filename.lower(), (int(sprite.w), int(sprite.h)))
    except Exception:
        # Atlas is a refinement source only; exporter must continue without it.
        atlas_size_by_png = {}

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  Error: openpyxl is required to read enriched workbook")
        return []

    def _clean_png_name(raw: object) -> str | None:
        if not isinstance(raw, str):
            return None
        candidate = raw.split(",", 1)[0].strip()
        if not candidate:
            return None
        return candidate

    wb = load_workbook(enriched_path, read_only=True, data_only=True)
    if "MapChip" not in wb.sheetnames:
        print("  Error: MapChip sheet missing in enriched workbook")
        return []

    ws = wb["MapChip"]
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {str(h).strip().lower(): i for i, h in enumerate(headers) if h is not None}
    required = ["id", "name", "res", "resfolder", "img", "pngname"]
    if any(key not in idx for key in required):
        print("  Error: enriched MapChip headers missing one of: id, name, resFolder, img, pngName")
        return []

    furniture_rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        row_id = row[idx["id"]]
        name = row[idx["name"]]
        res_value = row[idx["res"]]
        res_folder = row[idx["resfolder"]]
        img_id = row[idx["img"]]
        png_name = _clean_png_name(row[idx["pngname"]])

        if (
            row_id is None
            or not isinstance(name, str)
            or str(res_folder).lower() != "furniture"
            or res_value != 10
        ):
            continue

        if png_name is None and isinstance(img_id, int):
            png_name = img_inf.get(img_id)

        if not png_name:
            continue

        furniture_rows.append({
            "id": int(row_id),
            "name": name,
            "resFolder": "furniture",
            "img": int(img_id) if isinstance(img_id, int) else img_id,
            "pngName": png_name,
        })

    exported = []
    errors = []
    seen_names: set[str] = set()

    for item in furniture_rows:
        item_id = int(item["id"])
        item_name = item["name"]
        if item_name in seen_names:
            continue
        seen_names.add(item_name)

        name = item_name
        png_name = item["pngName"]
        png_path = furniture_assets_dir / png_name

        try:
            if not png_path.exists():
                errors.append(f"Furniture {name}: missing {png_name}")
                continue

            raw = Image.open(png_path).convert("RGBA")

            def _tight_crop(image: Image.Image) -> Image.Image:
                bbox = image.getbbox()
                if not bbox:
                    return image
                pad = 1
                left = max(0, bbox[0] - pad)
                top = max(0, bbox[1] - pad)
                right = min(image.width, bbox[2] + pad)
                bottom = min(image.height, bbox[3] + pad)
                return image.crop((left, top, right, bottom))

            def _fit_to_target(image: Image.Image, target_w: int, target_h: int) -> Image.Image:
                """Bottom-center fit image to target bounds (crop or pad)."""
                if target_w <= 0 or target_h <= 0:
                    return image

                cur = image
                # Crop if larger than target.
                if cur.width > target_w or cur.height > target_h:
                    left = max(0, (cur.width - target_w) // 2)
                    top = max(0, cur.height - target_h)
                    cur = cur.crop((left, top, min(cur.width, left + target_w), min(cur.height, top + target_h)))

                # Pad if smaller than target.
                if cur.width < target_w or cur.height < target_h:
                    canvas = Image.new("RGBA", (target_w, target_h), (0, 0, 0, 0))
                    x = max(0, (target_w - cur.width) // 2)
                    y = max(0, target_h - cur.height)
                    canvas.paste(cur, (x, y))
                    cur = canvas

                return cur

            def _keep_primary_component(image: Image.Image) -> Image.Image:
                """Keep only the dominant connected component to avoid clipped multi-frame bleed."""
                alpha = image.split()[3]
                w, h = alpha.size
                pix = alpha.load()
                seen = [[False] * w for _ in range(h)]
                components: list[dict] = []

                for y in range(h):
                    for x in range(w):
                        if seen[y][x] or pix[x, y] == 0:
                            continue

                        stack = [(x, y)]
                        seen[y][x] = True
                        area = 0
                        min_x = x
                        max_x = x
                        min_y = y
                        max_y = y
                        points: list[tuple[int, int]] = []

                        while stack:
                            cx, cy = stack.pop()
                            area += 1
                            points.append((cx, cy))
                            min_x = min(min_x, cx)
                            max_x = max(max_x, cx)
                            min_y = min(min_y, cy)
                            max_y = max(max_y, cy)

                            for nx, ny in ((cx + 1, cy), (cx - 1, cy), (cx, cy + 1), (cx, cy - 1)):
                                if 0 <= nx < w and 0 <= ny < h and (not seen[ny][nx]) and pix[nx, ny] > 0:
                                    seen[ny][nx] = True
                                    stack.append((nx, ny))

                        components.append(
                            {
                                "area": area,
                                "min_y": min_y,
                                "max_y": max_y,
                                "points": points,
                            }
                        )

                if len(components) < 2:
                    return image

                components.sort(key=lambda c: c["area"], reverse=True)
                primary = components[0]
                keep_points = set(primary["points"])

                cleaned = image.copy()
                out_pix = cleaned.load()
                for y in range(h):
                    for x in range(w):
                        if out_pix[x, y][3] == 0:
                            continue
                        if (x, y) not in keep_points:
                            out_pix[x, y] = (0, 0, 0, 0)

                return cleaned

            safe_name = name.replace(" ", "_").replace("/", "_").replace("'", "")
            normalized_name = " ".join(name.lower().split())
            source_stem = Path(png_name).stem.replace(" ", "_").replace("/", "_")

            # Build per-form variants from .opt when available (main form + numbered forms).
            variants: list[dict] = []
            opt_path = png_path.with_suffix(".opt")
            if opt_path.exists():
                try:
                    from parsers.opt_parser import parse_opt

                    opt_data = parse_opt(opt_path)
                    cell_w = int(opt_data.get("cell_width", 0) or 0)
                    cell_h = int(opt_data.get("cell_height", 0) or 0)
                    sprites = [
                        s
                        for s in opt_data.get("sprites", [])
                        if int(s.get("w", 0) or 0) > 0 and int(s.get("h", 0) or 0) > 0
                    ]

                    # Use row-major sequence order (frame 0,1,2,...) so main form is stable.
                    cols = int(opt_data.get("cols", 0) or 0)

                    def _seq_index(sprite: dict) -> int:
                        return int(sprite.get("v", 0) or 0) * cols + int(sprite.get("u", 0) or 0)

                    sprites = sorted(sprites, key=_seq_index)

                    atlas_target = atlas_size_by_png.get(png_name.lower())

                    for idx, sprite in enumerate(sprites, start=1):
                        if idx >= 2 and normalized_name in FURNITURE_PRIMARY_ONLY_VARIANTS:
                            continue

                        w = int(sprite["w"])
                        h = int(sprite["h"])
                        src_x = int(sprite.get("src_x", sprite.get("x", 0)))
                        src_y = int(sprite.get("src_y", sprite.get("y", 0)))
                        dest_x = int(sprite.get("dest_x", 0))
                        dest_y = int(sprite.get("dest_y", 0))
                        canvas_w = max(cell_w, dest_x + w, w)
                        canvas_h = max(cell_h, dest_y + h, h)

                        frame = Image.new("RGBA", (canvas_w, canvas_h), (0, 0, 0, 0))
                        cut = raw.crop((src_x, src_y, src_x + w, src_y + h))
                        frame.paste(cut, (dest_x, dest_y))
                        frame = _tight_crop(frame)

                        # Alternate forms can contain tiny detached pixels from neighboring atlas content.
                        if idx >= 2:
                            frame = _keep_primary_component(frame)
                            frame = _tight_crop(frame)

                        # Atlas envelope stabilizes the primary icon; keep alternates tightly cropped.
                        if idx == 1 and atlas_target:
                            frame = _fit_to_target(frame, atlas_target[0], atlas_target[1])

                        if scale > 1:
                            frame = frame.resize((frame.width * scale, frame.height * scale), Image.Resampling.NEAREST)

                        if idx == 1:
                            variant_filename = f"furniture_{item_id:03d}_{safe_name}__{source_stem}.png"
                        else:
                            variant_filename = f"furniture_{item_id:03d}_{safe_name}__{source_stem}_{idx}.png"

                        frame.save(furniture_dir / variant_filename, "PNG")
                        variants.append(
                            {
                                "index": idx,
                                "filename": variant_filename,
                                "u": int(sprite.get("u", -1)),
                                "v": int(sprite.get("v", -1)),
                                "w": w,
                                "h": h,
                            }
                        )
                except Exception as opt_error:
                    errors.append(f"Furniture {name}: opt parse failed ({opt_error})")

            if variants:
                filename = variants[0]["filename"]
            else:
                icon = _tight_crop(raw)
                if scale > 1:
                    icon = icon.resize((icon.width * scale, icon.height * scale), Image.Resampling.NEAREST)
                filename = f"furniture_{item_id:03d}_{safe_name}__{source_stem}.png"
                icon.save(furniture_dir / filename, "PNG")
                variants = [{"index": 1, "filename": filename}]

            exported.append({
                "id": item_id,
                "name": name,
                "filename": filename,
                "resFolder": item["resFolder"],
                "img": item["img"],
                "pngName": png_name,
                "variantCount": len(variants),
                "variants": variants,
            })

        except Exception as e:
            errors.append(f"Furniture {name}: {e}")

    print(f"  Exported: {len(exported)} furniture icons")
    if errors:
        print(f"  Errors: {len(errors)}")
        for err in errors[:5]:
            print(f"    {err}")

    return exported


def export_requested_icons(output_dir: Path) -> list[dict]:
    """Export user-requested icons by png name, sourced from original KA_assets files."""
    print("\n[Requested Icons]")

    if not REQUESTED_ICONS_DIR.exists():
        print(f"  Requested folder missing: {REQUESTED_ICONS_DIR}")
        return []

    requested_dir = output_dir / "requested"
    requested_dir.mkdir(parents=True, exist_ok=True)

    KA = config.KA_ASSETS_DIR
    if not KA.exists():
        print("  Error: KA_assets folder not found")
        return []

    # Index all PNG candidates by filename to resolve copies back to original assets.
    index: dict[str, list[Path]] = {}
    for candidate in KA.rglob("*.png"):
        index.setdefault(candidate.name.lower(), []).append(candidate)

    def _pick_preferred(paths: list[Path], wanted_name: str) -> Path:
        forced_rel = REQUESTED_SOURCE_OVERRIDES.get(wanted_name.lower())
        if forced_rel:
            forced = KA / forced_rel
            if forced in paths:
                return forced

        # Prefer non-localized root assets over language-specific variants.
        ranked = sorted(
            paths,
            key=lambda p: (
                "english.lproj" in str(p).lower()
                or "ko" in p.parts
                or "zh" in p.parts
                or "zh-cn" in p.parts,
                len(p.parts),
            ),
        )
        return ranked[0]

    exported: list[dict] = []
    errors: list[str] = []

    for wanted in sorted(REQUESTED_ICONS_DIR.glob("*.png"), key=lambda p: p.name.lower()):
        key = wanted.name.lower()
        matches = index.get(key, [])
        if not matches:
            errors.append(f"{wanted.name}: no matching original in KA_assets")
            continue

        source_path = _pick_preferred(matches, wanted.name)
        dest_path = requested_dir / wanted.name
        shutil.copy2(source_path, dest_path)

        exported.append(
            {
                "id": wanted.stem,
                "name": wanted.stem.replace("_", " "),
                "filename": wanted.name,
                "path": f"requested/{wanted.name}",
                "source": str(source_path.relative_to(KA)).replace("\\", "/"),
            }
        )

    print(f"  Exported: {len(exported)} requested icons")
    if errors:
        print(f"  Errors: {len(errors)}")
        for err in errors[:10]:
            print(f"    {err}")

    return exported


def main():
    parser = argparse.ArgumentParser(description="Export game icons for website")
    parser.add_argument("--output", "-o", type=Path, default=Path("website_icons"), 
                        help="Output directory for icons (default: website_icons/)")
    parser.add_argument("--scale", "-s", type=int, default=1, 
                        help="Scale factor for icons (default: 1)")
    args = parser.parse_args()
    
    output_dir = args.output
    scale = args.scale
    
    print("=" * 80)
    print("Exporting Kingdom Adventures Icons for Website")
    print("=" * 80)
    print(f"Output: {output_dir.absolute()}")
    print(f"Scale: {scale}x")
    
    # Export each category
    items_data = export_item_icons(output_dir, scale)
    equip_data = export_equipment_icons(output_dir, scale)
    eggs_data = export_egg_icons(output_dir, scale)
    attr_data = export_attribute_icons(output_dir, scale)
    gender_data = export_gender_icons(output_dir, scale)
    furniture_data = export_furniture_icons(output_dir, scale)
    requested_data = export_requested_icons(output_dir)

    print("\n[Quality Pass]")
    enhanced_count = enhance_exported_pngs(output_dir)
    print(f"  Alpha-bleed applied: {enhanced_count} PNG files")
    
    # Generate master manifest
    manifest = {
        "version": "1.0",
        "scale": scale,
        "items": items_data,
        "equipment": equip_data,
        "eggs": eggs_data,
        "attributes": attr_data,
        "gender": gender_data,
        "furniture": furniture_data,
        "requested": requested_data,
        "summary": {
            "items": len(items_data),
            "equipment": len(equip_data),
            "eggs": len(eggs_data),
            "attributes": len(attr_data),
            "gender": len(gender_data),
            "furniture": len(furniture_data),
            "requested": len(requested_data),
        }
    }
    
    manifest_path = output_dir / "manifest.json"
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)
    
    print(f"\n[Manifest] {manifest_path}")
    print(f"  Total icons exported: {manifest['summary']['items'] + manifest['summary']['equipment'] + manifest['summary']['eggs'] + manifest['summary']['attributes'] + manifest['summary']['gender']}")
    
    # Generate usage guide
    readme_path = output_dir / "README.md"
    with open(readme_path, "w", encoding="utf-8") as f:
        f.write(f"""# Kingdom Adventures Website Icons

Exported on: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
Scale: {scale}x

## Structure

```
{output_dir}/
  items/           {len(items_data)} item icons
  equipment/       {len(equip_data)} equipment icons
  eggs/            {len(eggs_data)} egg icons
  attributes/      {len(attr_data)} field attribute icons
  manifest.json    Complete metadata for all icons
```

## Usage

Each icon PNG is linked to its game entity through `manifest.json`.

### Items
- File: `items/item_XXX.png` (XXX = item ID with leading zeros)
- IDs 0-6: Material resources (Diamonds, Grass, Wood, Food, Ore, Mystic Ore, Energy)
- IDs 26-70: Localized items (Recovery Potions, Holy Herb, Sturdy Board, etc.)
- IDs 71+: Goods and materials

### Equipment
- File: `equipment/equip_XXX.png`
- Linked to equipment ID, type, and attribute
- Attributes: Ground(1), Grass(2), Sand(3), Rock(4), Volcano(5), Snow(6), Swamp(7)

### Eggs
- File: `eggs/egg_X.png` (X = 0-7)
- White, Blue, Green, Red, Purple, Black, Yellow, Rainbow

### Attributes
- File: `attributes/attribute_X_name.png` (X = 1-7)
- Ground, Grass, Sand, Rock, Volcano, Snow, Swamp

## Integration Example

```javascript
// Load manifest
const manifest = await fetch('/icons/manifest.json').then(r => r.json());

// Find item icon
const woodItem = manifest.items.find(i => i.name === "Wood");
console.log(woodItem.filename); // "item_002.png"

// Display icon
<img src="/icons/items/{{{{ woodItem.filename }}}}" alt="Wood" />

// Find equipment with Grass attribute
const grassWeapons = manifest.equipment.filter(e => e.attribute === 2);
```

## Replace Emoji Icons

Current website uses emoji for:
- 🪵 Wood → `items/item_002.png`
- 🪨 Ore → `items/item_004.png`
- 💎 Mystic Ore → `items/item_005.png`
- 🪙 Copper Coin → TBD (need to find coin icons in assets)
- ⚡ Energy → `items/item_006.png`

See manifest.json for complete ID→name→filename mapping.
""")
    
    print(f"\n[README] {readme_path}")
    print("\n" + "=" * 80)
    print("Export complete! Check the output directory for:")
    print(f"  - Icon PNGs organized by type")
    print(f"  - manifest.json with complete metadata")
    print(f"  - README.md with usage examples")
    print("=" * 80)


if __name__ == "__main__":
    main()
